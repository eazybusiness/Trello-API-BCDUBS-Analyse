"""
Script to calculate wages (Gagen) for voice actors based on Google Sheets script data.

This script:
1. Checks if gagen.xlsx already exists as attachment in Trello list "Skripte zur Aufnahme"
2. If not, downloads and analyzes the Google Sheets from the card description
3. Calculates wages based on lines per speaker and character groups
4. Creates a gagen.xlsx file locally (read-only mode to protect production data)

Usage:
    python calculate_gagen.py
"""

import os
import re
import json
import requests
import pandas as pd
from io import BytesIO
from typing import Dict, List, Tuple, Optional
from trello_client import TrelloClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ColorMatcher:
    """Helper class to match colors with tolerance."""
    
    # Character group color definitions (RGB hex values)
    COLOR_GROUPS = {
        'Zivis': '4f81bd',
        'Cops': 'c0504d',
        'Weibliche Sprecher': 'f79646',
        'Erzähler': '9bbb59'
    }
    
    @staticmethod
    def hex_to_rgb(hex_color: str) -> Tuple[int, int, int]:
        """Convert hex color to RGB tuple."""
        hex_color = hex_color.lstrip('#').lower()
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    @staticmethod
    def color_distance(color1: str, color2: str) -> float:
        """Calculate Euclidean distance between two colors."""
        rgb1 = ColorMatcher.hex_to_rgb(color1)
        rgb2 = ColorMatcher.hex_to_rgb(color2)
        return sum((a - b) ** 2 for a, b in zip(rgb1, rgb2)) ** 0.5
    
    @staticmethod
    def match_color_group(color: str, tolerance: float = 30.0) -> Optional[str]:
        """
        Match a color to a character group with tolerance.
        
        Args:
            color: Hex color string
            tolerance: Maximum color distance to consider a match (default: 30)
        
        Returns:
            Character group name or None if no match
        """
        if not color:
            return None
        
        color = color.lstrip('#').lower()
        best_match = None
        best_distance = float('inf')
        
        for group_name, group_color in ColorMatcher.COLOR_GROUPS.items():
            distance = ColorMatcher.color_distance(color, group_color)
            if distance < best_distance and distance <= tolerance:
                best_distance = distance
                best_match = group_name
        
        return best_match


class GagenCalculator:
    """Main class for calculating wages from Google Sheets data."""
    
    def __init__(self):
        self.trello_client = TrelloClient()
        self.board_name = "True Crime Video Dubs"
        self.list_name = "Skripte zur Aufnahme"
    
    def find_card_in_list(self, list_name: str) -> Optional[Dict]:
        """
        Find cards in a specific list.
        
        Args:
            list_name: Name of the Trello list
        
        Returns:
            List of cards in the specified list
        """
        result = self.trello_client.get_board_cards_with_lists(self.board_name)
        cards = result['cards_by_list'].get(list_name, [])
        return cards
    
    def check_attachment_exists(self, card: Dict, filename: str) -> bool:
        """
        Check if a specific attachment exists on a card.
        
        Args:
            card: Trello card data
            filename: Name of the attachment to check
        
        Returns:
            True if attachment exists, False otherwise
        """
        # Get attachments from card
        card_id = card['id']
        url = f"{self.trello_client.base_url}/cards/{card_id}/attachments"
        params = {
            'key': self.trello_client.api_key,
            'token': self.trello_client.token
        }
        
        response = requests.get(url, params=params)
        response.raise_for_status()
        attachments = response.json()
        
        for attachment in attachments:
            if attachment.get('name', '').lower() == filename.lower():
                return True
        
        return False
    
    def extract_google_sheets_url(self, description: str) -> Optional[str]:
        """
        Extract Google Sheets URL from card description.
        
        Args:
            description: Card description text
        
        Returns:
            Google Sheets URL or None
        """
        pattern = r'https://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)'
        match = re.search(pattern, description)
        if match:
            spreadsheet_id = match.group(1)
            return f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}'
        return None
    
    def download_google_sheet(self, sheet_url: str) -> pd.DataFrame:
        """
        Download Google Sheet as CSV and parse into DataFrame.
        
        Args:
            sheet_url: Google Sheets URL
        
        Returns:
            DataFrame with sheet data
        """
        # Extract spreadsheet ID
        pattern = r'/d/([a-zA-Z0-9-_]+)'
        match = re.search(pattern, sheet_url)
        if not match:
            raise ValueError(f"Invalid Google Sheets URL: {sheet_url}")
        
        spreadsheet_id = match.group(1)
        
        # Try CSV export (works for public sheets)
        csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv"
        
        print(f"Downloading sheet from: {csv_url}")
        response = requests.get(csv_url, timeout=10)
        
        if response.status_code == 200:
            # Parse CSV
            df = pd.read_csv(BytesIO(response.content))
            return df
        elif response.status_code in [401, 403]:
            raise PermissionError("Google Sheet is not publicly accessible. Please set sharing to 'Anyone with link can view'")
        else:
            raise Exception(f"Failed to download sheet. Status code: {response.status_code}")
    
    def parse_timecode(self, timecode: str) -> float:
        """
        Parse timecode in format HH:MM:SS:FF to total minutes.
        
        Args:
            timecode: Timecode string (e.g., "00:05:23:12")
        
        Returns:
            Total minutes as float
        """
        if not timecode or pd.isna(timecode):
            return 0.0
        
        try:
            # Handle format HH:MM:SS:FF or HH:MM:SS
            parts = str(timecode).strip().split(':')
            if len(parts) >= 3:
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                # Ignore frames (FF) if present
                
                total_minutes = hours * 60 + minutes + seconds / 60.0
                return total_minutes
            else:
                return 0.0
        except (ValueError, IndexError):
            return 0.0
    
    def analyze_script_data(self, df: pd.DataFrame) -> Dict:
        """
        Analyze script data from DataFrame.
        
        Args:
            df: DataFrame with script data
        
        Returns:
            Dictionary with analysis results
        """
        # The actual data structure is:
        # Row 0: Header with speaker names
        # Row 1: Column headers (Voices, Paragraph index, Speaker, Start time, End time, German, Text, etc.)
        # Row 2+: Data rows
        
        # Find the header row (contains "Voices", "Paragraph index", etc.)
        header_row = None
        for i in range(min(5, len(df))):
            row_values = df.iloc[i].astype(str).str.lower()
            if any('voices' in str(val).lower() for val in row_values):
                header_row = i
                break
        
        if header_row is None:
            print("⚠️  WARNING: Could not find header row. Using default structure.")
            header_row = 1
        
        # Re-read with correct header
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        df.columns = ['Voices', 'Paragraph index', 'Speaker', 'Start time', 'End time', 'German', 'Text', 'Notes'][:len(df.columns)]
        
        # Get total minutes from last End time
        if 'End time' in df.columns:
            end_times = df['End time'].dropna()
            if len(end_times) > 0:
                last_timecode = end_times.iloc[-1]
                total_minutes = self.parse_timecode(last_timecode)
                # Round up to next full minute
                total_minutes = int(total_minutes) + (1 if total_minutes % 1 > 0 else 0)
            else:
                total_minutes = 0
        else:
            total_minutes = 0
        
        # Get total lines from Paragraph index
        if 'Paragraph index' in df.columns:
            line_numbers = pd.to_numeric(df['Paragraph index'], errors='coerce').dropna()
            total_lines = int(line_numbers.max()) if len(line_numbers) > 0 else len(df)
        else:
            total_lines = len(df)
        
        # Count lines per speaker
        lines_per_speaker = {}
        lines_per_color_group = {group: 0 for group in ColorMatcher.COLOR_GROUPS.keys()}
        
        if 'Speaker' in df.columns:
            speaker_data = df['Speaker'].dropna()
            speaker_counts = speaker_data.value_counts()
            lines_per_speaker = speaker_counts.to_dict()
            
            # Note: We cannot extract background colors from CSV
            print("\n⚠️  WARNING: Background colors cannot be extracted from CSV export.")
            print("    Color-based character group analysis will be skipped.")
            print("    To get color data, you would need to use Google Sheets API with authentication.")
        
        return {
            'total_minutes': total_minutes,
            'total_lines': total_lines,
            'lines_per_speaker': lines_per_speaker,
            'lines_per_color_group': lines_per_color_group
        }
    
    def check_express_label(self, card: Dict) -> bool:
        """
        Check if card has 'Express' label.
        
        Args:
            card: Trello card data
        
        Returns:
            True if Express label exists
        """
        labels = card.get('labels', [])
        for label in labels:
            if label.get('name', '').lower() == 'express':
                return True
        return False
    
    def calculate_wages(self, analysis: Dict, is_express: bool) -> Dict:
        """
        Calculate wages based on analysis data.
        
        Args:
            analysis: Analysis results from analyze_script_data
            is_express: Whether this is an express project
        
        Returns:
            Dictionary with wage calculations
        """
        rate_per_line = 8.75 if not is_express else 9.75
        total_minutes = analysis['total_minutes']
        total_lines = analysis['total_lines']
        
        if total_lines == 0:
            print("⚠️  WARNING: Total lines is 0. Cannot calculate wages.")
            return {
                'rate_per_line': rate_per_line,
                'base_rate': 0,
                'speaker_wages': {},
                'color_group_wages': {}
            }
        
        # Base rate: rate * minutes / total_lines
        base_rate = rate_per_line * total_minutes / total_lines
        
        # Calculate per speaker
        speaker_wages = {}
        for speaker, lines in analysis['lines_per_speaker'].items():
            wage = base_rate * lines
            speaker_wages[speaker] = {
                'lines': lines,
                'wage': round(wage, 2)
            }
        
        # Calculate per color group
        color_group_wages = {}
        for group, lines in analysis['lines_per_color_group'].items():
            if lines > 0:
                wage = base_rate * lines
                color_group_wages[group] = {
                    'lines': lines,
                    'wage': round(wage, 2)
                }
        
        return {
            'rate_per_line': rate_per_line,
            'base_rate': round(base_rate, 4),
            'total_minutes': total_minutes,
            'total_lines': total_lines,
            'speaker_wages': speaker_wages,
            'color_group_wages': color_group_wages
        }
    
    def create_excel_report(self, card_name: str, wages: Dict, output_file: str = 'gagen.xlsx'):
        """
        Create Excel report with wage calculations.
        
        Args:
            card_name: Name of the Trello card
            wages: Wage calculation results
            output_file: Output filename
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet
        ws_summary = wb.create_sheet('Zusammenfassung')
        
        # Header
        ws_summary['A1'] = 'Gagen-Berechnung'
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f'Projekt: {card_name}'
        ws_summary['A2'].font = Font(size=12, italic=True)
        
        # Calculation parameters
        ws_summary['A4'] = 'Berechnungsgrundlage:'
        ws_summary['A4'].font = Font(bold=True)
        ws_summary['A5'] = 'Satz pro Line:'
        ws_summary['B5'] = f"€{wages['rate_per_line']}"
        ws_summary['A6'] = 'Gesamtminuten:'
        ws_summary['B6'] = wages['total_minutes']
        ws_summary['A7'] = 'Gesamtanzahl Lines:'
        ws_summary['B7'] = wages['total_lines']
        ws_summary['A8'] = 'Basis-Rate (€/Line):'
        ws_summary['B8'] = f"€{wages['base_rate']}"
        
        # Speaker wages
        ws_speakers = wb.create_sheet('Sprecher')
        ws_speakers['A1'] = 'Sprecher'
        ws_speakers['B1'] = 'Anzahl Lines'
        ws_speakers['C1'] = 'Gage (€)'
        
        # Style header
        for cell in ['A1', 'B1', 'C1']:
            ws_speakers[cell].font = Font(bold=True)
            ws_speakers[cell].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            ws_speakers[cell].font = Font(bold=True, color='FFFFFF')
        
        row = 2
        total_speaker_wage = 0
        for speaker, data in sorted(wages['speaker_wages'].items()):
            ws_speakers[f'A{row}'] = speaker
            ws_speakers[f'B{row}'] = data['lines']
            ws_speakers[f'C{row}'] = data['wage']
            ws_speakers[f'C{row}'].number_format = '€#,##0.00'
            total_speaker_wage += data['wage']
            row += 1
        
        # Total row
        ws_speakers[f'A{row}'] = 'GESAMT'
        ws_speakers[f'A{row}'].font = Font(bold=True)
        ws_speakers[f'C{row}'] = total_speaker_wage
        ws_speakers[f'C{row}'].number_format = '€#,##0.00'
        ws_speakers[f'C{row}'].font = Font(bold=True)
        
        # Color group wages (if available)
        if wages['color_group_wages']:
            ws_groups = wb.create_sheet('Charaktergruppen')
            ws_groups['A1'] = 'Charaktergruppe'
            ws_groups['B1'] = 'Anzahl Lines'
            ws_groups['C1'] = 'Gage (€)'
            
            # Style header
            for cell in ['A1', 'B1', 'C1']:
                ws_groups[cell].font = Font(bold=True)
                ws_groups[cell].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                ws_groups[cell].font = Font(bold=True, color='FFFFFF')
            
            row = 2
            total_group_wage = 0
            for group, data in sorted(wages['color_group_wages'].items()):
                ws_groups[f'A{row}'] = group
                ws_groups[f'B{row}'] = data['lines']
                ws_groups[f'C{row}'] = data['wage']
                ws_groups[f'C{row}'].number_format = '€#,##0.00'
                total_group_wage += data['wage']
                row += 1
            
            # Total row
            ws_groups[f'A{row}'] = 'GESAMT'
            ws_groups[f'A{row}'].font = Font(bold=True)
            ws_groups[f'C{row}'] = total_group_wage
            ws_groups[f'C{row}'].number_format = '€#,##0.00'
            ws_groups[f'C{row}'].font = Font(bold=True)
        
        # Adjust column widths
        for ws in wb.worksheets:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
        
        # Save workbook
        wb.save(output_file)
        print(f"\n✅ Excel report created: {output_file}")
    
    def run(self):
        """Main execution method."""
        print("="*70)
        print("GAGEN-BERECHNUNG (Wage Calculation)")
        print("="*70)
        
        # Find cards in "Skripte zur Aufnahme" list
        print(f"\n1. Searching for cards in list '{self.list_name}'...")
        cards = self.find_card_in_list(self.list_name)
        
        if not cards:
            print(f"❌ No cards found in list '{self.list_name}'")
            return
        
        print(f"✅ Found {len(cards)} card(s)")
        
        # Process each card
        for i, card in enumerate(cards, 1):
            print(f"\n{'='*70}")
            print(f"Processing card {i}/{len(cards)}: {card['name']}")
            print(f"{'='*70}")
            
            # Check if gagen.xlsx already exists
            print("\n2. Checking for existing gagen.xlsx attachment...")
            if self.check_attachment_exists(card, 'gagen.xlsx'):
                print("✅ gagen.xlsx already exists as attachment. Skipping this card.")
                continue
            else:
                print("ℹ️  gagen.xlsx not found. Proceeding with calculation.")
            
            # Extract Google Sheets URL
            print("\n3. Extracting Google Sheets URL from description...")
            sheet_url = self.extract_google_sheets_url(card.get('desc', ''))
            
            if not sheet_url:
                print("❌ No Google Sheets URL found in card description. Skipping.")
                continue
            
            print(f"✅ Found Google Sheets URL: {sheet_url}")
            
            # Download and analyze
            print("\n4. Downloading and analyzing Google Sheet...")
            try:
                df = self.download_google_sheet(sheet_url)
                print(f"✅ Downloaded sheet with {len(df)} rows and {len(df.columns)} columns")
                
                analysis = self.analyze_script_data(df)
                print(f"\n📊 Analysis Results:")
                print(f"   Total Minutes: {analysis['total_minutes']}")
                print(f"   Total Lines: {analysis['total_lines']}")
                print(f"   Speakers found: {len(analysis['lines_per_speaker'])}")
                
            except Exception as e:
                print(f"❌ Error analyzing sheet: {e}")
                continue
            
            # Check for Express label
            print("\n5. Checking for Express label...")
            is_express = self.check_express_label(card)
            if is_express:
                print("✅ Express label found. Using rate: €9.75")
            else:
                print("ℹ️  No Express label. Using standard rate: €8.75")
            
            # Calculate wages
            print("\n6. Calculating wages...")
            wages = self.calculate_wages(analysis, is_express)
            
            print(f"\n💰 Wage Calculation:")
            print(f"   Base rate: €{wages['base_rate']}/line")
            print(f"   Total speakers: {len(wages['speaker_wages'])}")
            
            # Create Excel report
            print("\n7. Creating Excel report...")
            output_file = f"gagen_{card['id']}.xlsx"
            self.create_excel_report(card['name'], wages, output_file)
            
            print(f"\n✅ Processing complete for card: {card['name']}")
            print(f"   Output file: {output_file}")
            print(f"\n⚠️  NOTE: File created locally only (read-only mode to protect production data)")
            print(f"   To upload to Trello, you would need write permissions.")


def main():
    """Main entry point."""
    try:
        calculator = GagenCalculator()
        calculator.run()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
